import mysql.connector
import openpyxl

#DADOS PARA A CONEXAO
fonte='localhost'
banco='world'
usuario='root'
senha='leojunior2020'

con = mysql.connector.connect(host=fonte,database = banco,user=usuario,password=senha)

consulta="select * from city"

book=openpyxl.Workbook()
#print(book.sheetnames)
book.create_sheet('Cidades')
city_page=book['Cidades']
city_page.append(['ID','NOME','DISTRITO','POPULACAO'])

cursor =  con.cursor()
cursor.execute(consulta)
linhas = cursor.fetchall()
#print("Numero total de registos: ", cursor.rowcount())

print("Mostrando os registos encontrados")
for linha in linhas:
    
    print(linha[1],"\n")
    city_page.append(linha)

#criacao do ficheiro
book.save('cidades2.xlsx')

if con.is_connected():
    cursor.close()
    con.close()
    print("CONEXAO AO MYSQL FOI ENCERRADA")
