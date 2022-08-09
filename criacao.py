import openpyxl

book=openpyxl.Workbook()
#print(book.sheetnames)
book.create_sheet('Frutas')

frutas_page=book['Frutas']
frutas_page.append(['PRODUTO','QUANTIDADE','PRECO'])
frutas_page.append(['banana',5,'R$11,60'])
frutas_page.append(['manga',121,'R$14,60'])
frutas_page.append(['maca',12,'R$15,60'])
frutas_page.append(['abacate',17,'R$87,60'])

book.save('compra.xlsx')